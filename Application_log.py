"""
Application Usage Dashboard
===========================

A simple, offline Streamlit dashboard for analysing log files produced by
several intranet applications.

UI
----
  * Application dropdown (8 applications)
  * Year dropdown
  * Month dropdown
  * Calculate button
  * Results table
  * Totals row at the bottom
  * Export to Excel

Design
------
Every application's calculation logic lives in its own function
(``calculate_app1`` ... ``calculate_app8``).  The UI never contains business
logic; it only looks up the selected application in ``APP_REGISTRY`` and calls
that application's function.  Adding or changing an application therefore only
means editing one function.

Each calculate function has the same signature::

    calculate_appN(base_path: str, year: int, month: int) -> (pandas.DataFrame, dict)

It returns the **daily** rows as a DataFrame (the first column is always the
date) and a ``meta`` dict that may carry extra headline numbers to show above
the table (used by App3 for whole-month distinct-user counts).

Run with::

    streamlit run app.py
"""

from __future__ import annotations

import calendar
import datetime as _dt
import os
import re
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _month_dates(year: int, month: int):
    """Return a list of date objects for every day in the given month."""
    n_days = calendar.monthrange(year, month)[1]
    return [_dt.date(year, month, d) for d in range(1, n_days + 1)]


def _count_lines_containing(filepath: str, needle: str) -> int:
    """
    Count how many *lines* in ``filepath`` contain ``needle``.

    This is line-based (Option B): a line that contains the keyword twice is
    still counted once.  The match is case-sensitive and is a plain substring
    test.  A missing file counts as zero.
    """
    if not os.path.isfile(filepath):
        return 0
    count = 0
    with open(filepath, "r", encoding="utf-8", errors="ignore") as fh:
        for line in fh:
            if needle in line:
                count += 1
    return count


def _iter_existing_day_files(base_path: str, year: int, month: int, name_builder):
    """
    Yield ``(date, filepath)`` for every day of the month whose log file
    actually exists on disk.  ``name_builder`` maps a date to a file name.

    Days with no log file are skipped so the table only shows days that ran.
    """
    if not base_path or not os.path.isdir(base_path):
        return
    for day in _month_dates(year, month):
        filepath = os.path.join(base_path, name_builder(day))
        if os.path.isfile(filepath):
            yield day, filepath


def _iso(day: _dt.date) -> str:
    return day.strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# App1 -- /home/dlpda/production_v3/chat_hist/
# File name: chat_hist_DD-MM-YYYY.log
# Usage = number of lines containing "thinking:"
# ---------------------------------------------------------------------------
def calculate_app1(base_path: str, year: int, month: int):
    name_builder = lambda d: f"chat_hist_{d.strftime('%d-%m-%Y')}.log"
    rows = []
    for day, filepath in _iter_existing_day_files(base_path, year, month, name_builder):
        rows.append({"Date": _iso(day),
                     "Usage": _count_lines_containing(filepath, "Thinking:")})
    df = pd.DataFrame(rows, columns=["Date", "Usage"])
    return df, {}


# ---------------------------------------------------------------------------
# App2 -- /home/dlpda/production_v3/chat_hist/
# File name: chat_hist_DD-MM-YYYY.log
# Usage = Count("Thinking:") - Count("chat_Query:")   (line-based, per day)
# ---------------------------------------------------------------------------
def calculate_app2(base_path: str, year: int, month: int):
    name_builder = lambda d: f"chat_hist_{d.strftime('%d-%m-%Y')}.log"
    rows = []
    for day, filepath in _iter_existing_day_files(base_path, year, month, name_builder):
        Thinking = _count_lines_containing(filepath, "Thinking:")
        chat_query = _count_lines_containing(filepath, "chat_Query:")
        rows.append({"Date": _iso(day), "Usage": Thinking - chat_query})
    df = pd.DataFrame(rows, columns=["Date", "Usage"])
    return df, {}


# ---------------------------------------------------------------------------
# App3 -- /home/dlpda/production_v3/mom_app_v2.0/logs/
# File name: YYYY-MM-DD.txt
#
# Two metrics per day:
#   * Unique Login Users -- distinct users with an "AD Login Success"
#   * Unique Usage Users -- users who logged in AND produced at least one
#     "Status: SUCCESS" (any feature counts, one per user)
#
# Most "Status: SUCCESS" lines have no User field (only an IP), so each
# success is attributed to whoever is currently active on that IP.  We scan
# the file top-to-bottom (already time-ordered) and keep an IP -> user map
# that any line carrying a User field refreshes.  User IDs are uppercased so
# the same person is never counted twice on a case difference.
# ---------------------------------------------------------------------------
_IP_RE = re.compile(r"IP:\s*(\S+)")
_USER_RE = re.compile(r"User:\s*(\S+)")


def _app3_parse_file(filepath: str):
    """Return (login_users, usage_users) as sets of upper-cased user IDs."""
    login_users: set[str] = set()
    success_users: set[str] = set()
    ip_to_user: dict[str, str] = {}

    with open(filepath, "r", encoding="utf-8", errors="ignore") as fh:
        for line in fh:
            ip_m = _IP_RE.search(line)
            ip = ip_m.group(1) if ip_m else None
            user_m = _USER_RE.search(line)
            user = user_m.group(1).upper() if user_m else None

            # Any line naming a user tells us who is active on that IP.
            if ip and user:
                ip_to_user[ip] = user

            if "AD Login Success" in line and user:
                login_users.add(user)

            if "Status: SUCCESS" in line:
                if user:                       # line names the user directly
                    success_users.add(user)
                elif ip and ip in ip_to_user:  # nameless success -> IP's user
                    success_users.add(ip_to_user[ip])

    # A usage user must also have logged in.
    usage_users = success_users & login_users
    return login_users, usage_users


def calculate_app3(base_path: str, year: int, month: int):
    name_builder = lambda d: f"{d.strftime('%Y-%m-%d')}.txt"
    rows = []
    month_login: set[str] = set()
    month_usage: set[str] = set()

    for day, filepath in _iter_existing_day_files(base_path, year, month, name_builder):
        login_users, usage_users = _app3_parse_file(filepath)
        rows.append({
            "Date": _iso(day),
            "Unique Login Users": len(login_users),
            "Unique Usage Users": len(usage_users),
        })
        month_login |= login_users
        month_usage |= usage_users

    df = pd.DataFrame(rows, columns=["Date", "Unique Login Users", "Unique Usage Users"])
    meta = {
        "extra_metrics": {
            "Distinct login users (whole month)": len(month_login),
            "Distinct usage users (whole month)": len(month_usage),
        }
    }
    return df, meta


# ---------------------------------------------------------------------------
# App4 -- /home/dlpda/production_v3/Document_Asst/logs/
# File name: activity_YYYY-MM-DD.log
#   Users     = number of lines containing "TURN:1"
#   Questions = number of lines containing "QUESTION:"
# ---------------------------------------------------------------------------
def calculate_app4(base_path: str, year: int, month: int):
    name_builder = lambda d: f"activity_{d.strftime('%Y-%m-%d')}.log"
    rows = []
    for day, filepath in _iter_existing_day_files(base_path, year, month, name_builder):
        rows.append({
            "Date": _iso(day),
            "Users": _count_lines_containing(filepath, "TURN:1"),
            "Questions": _count_lines_containing(filepath, "QUESTION:"),
        })
    df = pd.DataFrame(rows, columns=["Date", "Users", "Questions"])
    return df, {}


# ---------------------------------------------------------------------------
# App5 -- RESERVED.  Placeholder only; implement later.
# ---------------------------------------------------------------------------
def calculate_app5(base_path: str, year: int, month: int):
    raise NotImplementedError(
        "App5 is reserved and has not been implemented yet."
    )


# ---------------------------------------------------------------------------
# App6 / App7 / App8 -- chat query counters
# File name: chat_history_YYYY-MM-DD.log
# Usage = number of lines containing "Query:"
# Only the base path differs between them, so they share one helper while
# still exposing a separate named function each (per the modular requirement).
# ---------------------------------------------------------------------------
def _calculate_query_app(base_path: str, year: int, month: int):
    name_builder = lambda d: f"chat_history_{d.strftime('%Y-%m-%d')}.log"
    rows = []
    for day, filepath in _iter_existing_day_files(base_path, year, month, name_builder):
        rows.append({"Date": _iso(day),
                     "Usage": _count_lines_containing(filepath, "Query:")})
    df = pd.DataFrame(rows, columns=["Date", "Usage"])
    return df, {}


def calculate_app6(base_path: str, year: int, month: int):
    return _calculate_query_app(base_path, year, month)


def calculate_app7(base_path: str, year: int, month: int):
    return _calculate_query_app(base_path, year, month)


def calculate_app8(base_path: str, year: int, month: int):
    return _calculate_query_app(base_path, year, month)


# ---------------------------------------------------------------------------
# Registry: the single place the UI looks things up.
# ---------------------------------------------------------------------------
APP_REGISTRY = {
    "App1": {
        "path": "/home/dlpda/production_v3/chat_hist/",
        "func": calculate_app1,
        "desc": "Daily usage = lines containing 'Thinking:'.",
        "reserved": False,
    },
    "App2": {
        "path": "/home/dlpda/production_v3/chat_hist/",
        "func": calculate_app2,
        "desc": "Daily usage = Count('Thinking:') - Count('chat_Query:').",
        "reserved": False,
    },
    "App3": {
        "path": "/home/dlpda/production_v3/mom_app_v2.0/logs/",
        "func": calculate_app3,
        "desc": "Unique login users and unique usage users (logged in AND a Status: SUCCESS).",
        "reserved": False,
    },
    "App4": {
        "path": "/home/dlpda/production_v3/Document_Asst/logs/",
        "func": calculate_app4,
        "desc": "Users = lines with 'TURN:1'; Questions = lines with 'QUESTION:'.",
        "reserved": False,
    },
    "App5": {
        "path": None,
        "func": calculate_app5,
        "desc": "Reserved - not implemented yet.",
        "reserved": True,
    },
    "App6": {
        "path": "/home/dlpda/proc/chat_logs/",
        "func": calculate_app6,
        "desc": "Daily usage = lines containing 'Query:'.",
        "reserved": False,
    },
    "App7": {
        "path": "/home/dlpda/aero/chat_logs/",
        "func": calculate_app7,
        "desc": "Daily usage = lines containing 'Query:'.",
        "reserved": False,
    },
    "App8": {
        "path": "/home/dlpda/tmd/chat_logs/",
        "func": calculate_app8,
        "desc": "Daily usage = lines containing 'Query:'.",
        "reserved": False,
    },
}

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ---------------------------------------------------------------------------
# Totals + Excel export
# ---------------------------------------------------------------------------
def build_totals_row(df: pd.DataFrame) -> dict:
    """Return a totals row: first column labelled 'TOTAL', numeric columns summed."""
    if df.empty:
        return {}
    first_col = df.columns[0]
    numeric_cols = df.select_dtypes(include="number").columns
    totals = {}
    for col in df.columns:
        if col == first_col:
            totals[col] = "TOTAL"
        elif col in numeric_cols:
            totals[col] = int(df[col].sum())
        else:
            totals[col] = ""
    return totals


def to_excel_bytes(df: pd.DataFrame, totals: dict, extra_metrics: dict,
                   title: str, sheet_name: str = "Usage") -> bytes:
    """Render the daily table + totals row (+ extra metrics) into an .xlsx in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    def arial(**kw):
        return Font(name="Arial", **kw)

    header_fill = PatternFill("solid", fgColor="305496")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")

    # Title line
    ws.append([title])
    ws["A1"].font = arial(bold=True, size=14)
    ws.append([])  # spacer

    header_row = ws.max_row + 1
    ws.append(list(df.columns))
    for cell in ws[header_row]:
        cell.font = arial(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center

    # Data
    for _, row in df.iterrows():
        ws.append([row[c] for c in df.columns])

    # Totals
    if totals:
        ws.append([totals.get(c, "") for c in df.columns])
        for cell in ws[ws.max_row]:
            cell.font = arial(bold=True)
            cell.fill = total_fill

    # Make the whole grid Arial
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
        for cell in row:
            if cell.font is None or cell.font.name != "Arial":
                bold = bool(cell.font and cell.font.bold)
                color = cell.font.color if cell.font else None
                cell.font = Font(name="Arial", bold=bold, color=color)

    # Extra headline metrics (App3)
    if extra_metrics:
        ws.append([])
        for key, value in extra_metrics.items():
            ws.append([key, value])
            r = ws.max_row
            ws.cell(row=r, column=1).font = arial(bold=True)
            ws.cell(row=r, column=2).font = arial()

    # Column widths
    for idx, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), 12) + 4
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
def main():
    st.set_page_config(page_title="Application Usage Dashboard", layout="centered")
    st.title("Application Usage Dashboard")
    st.caption("Analyse daily usage from offline intranet application logs.")

    this_year = _dt.date.today().year
    years = list(range(this_year - 3, this_year + 2))

    # --- Controls ---
    # Row 1: Application (full width)
    app_choice = st.selectbox("Application", list(APP_REGISTRY.keys()))

    # Row 2: Month and Year side by side
    col_month, col_year = st.columns(2)
    with col_month:
        month_name = st.selectbox("Month", MONTHS, index=_dt.date.today().month - 1)
    with col_year:
        year = st.selectbox("Year", years, index=years.index(this_year))
    month = MONTHS.index(month_name) + 1

    info = APP_REGISTRY[app_choice]
    st.caption(f"**{app_choice}** — {info['desc']}")
    if info["path"]:
        st.caption(f"Log folder: `{info['path']}`")

    calculate = st.button("Calculate", type="primary")

    # --- Reserved app short-circuit ---
    if info["reserved"]:
        st.info(f"{app_choice} is reserved and has not been implemented yet.")
        return

    # --- Run calculation on click; keep results in session_state ---
    if calculate:
        try:
            df, meta = info["func"](info["path"], year, month)
            st.session_state["result"] = {
                "app": app_choice,
                "year": year,
                "month_name": month_name,
                "df": df,
                "meta": meta,
            }
        except NotImplementedError as exc:
            st.info(str(exc))
            st.session_state.pop("result", None)
            return
        except Exception as exc:  # noqa: BLE001 - surface any parsing/IO issue plainly
            st.error(f"Could not calculate: {exc}")
            st.session_state.pop("result", None)
            return

    result = st.session_state.get("result")
    if not result:
        st.write("Choose an application, year and month, then press **Calculate**.")
        return

    df: pd.DataFrame = result["df"]
    meta: dict = result["meta"]
    title = f"{result['app']} usage — {result['month_name']} {result['year']}"

    st.subheader(title)

    if df.empty:
        folder = APP_REGISTRY[result["app"]]["path"]
        st.warning(
            "No log files were found for this application in the selected month.\n\n"
            f"Looked in: `{folder}`"
        )
        return

    # Extra headline metrics (App3 whole-month distinct users)
    extra_metrics = meta.get("extra_metrics", {})
    if extra_metrics:
        metric_cols = st.columns(len(extra_metrics))
        for col, (label, value) in zip(metric_cols, extra_metrics.items()):
            col.metric(label, value)

    # Table with a totals row appended at the bottom
    totals = build_totals_row(df)
    display_df = df.copy()
    if totals:
        display_df = pd.concat([display_df, pd.DataFrame([totals])], ignore_index=True)

    st.dataframe(display_df, use_container_width=True, hide_index=True)

    # Excel export
    xlsx = to_excel_bytes(df, totals, extra_metrics, title, sheet_name=result["app"])
    fname = f"{result['app']}_{result['year']}_{month:02d}_usage.xlsx"
    st.download_button(
        "Export to Excel",
        data=xlsx,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
